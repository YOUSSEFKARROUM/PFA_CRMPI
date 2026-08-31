"""
database.py — Couche d'accès aux données (SQLite) pour la plateforme CMRPI.

Fournit les opérations CRUD pour : PME (utilisateurs), audits, réponses,
domaines, questions et recommandations.
"""
import sqlite3
import uuid
import json
from datetime import datetime
from pathlib import Path
from contextlib import contextmanager

DB_PATH = Path(__file__).resolve().parent.parent / "app.db"
SCHEMA_PATH = Path(__file__).resolve().parent.parent / "data" / "schema.sql"


@contextmanager
def get_connection():
    """Fournit une connexion SQLite avec support des clés étrangères."""
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA foreign_keys = ON")
    conn.row_factory = sqlite3.Row
    try:
        yield conn
        conn.commit()
    finally:
        conn.close()


def init_database():
    """Initialise la base de données à partir du schéma SQL."""
    with get_connection() as conn:
        with open(SCHEMA_PATH, "r", encoding="utf-8") as f:
            conn.executescript(f.read())
    return True


def new_uuid() -> str:
    return str(uuid.uuid4())


# ============================================================
# PME (utilisateurs)
# ============================================================

def create_pme(email: str, password_hash: str, name: str, sector: str = None) -> str:
    pme_id = new_uuid()
    with get_connection() as conn:
        conn.execute(
            "INSERT INTO pme (id, email, password_hash, name, sector) VALUES (?, ?, ?, ?, ?)",
            (pme_id, email, password_hash, name, sector),
        )
    return pme_id


def get_pme_by_email(email: str):
    with get_connection() as conn:
        row = conn.execute("SELECT * FROM pme WHERE email = ?", (email,)).fetchone()
        return dict(row) if row else None


def get_pme_by_id(pme_id: str):
    with get_connection() as conn:
        row = conn.execute("SELECT * FROM pme WHERE id = ?", (pme_id,)).fetchone()
        return dict(row) if row else None


# ============================================================
# Questions & Domaines
# ============================================================

def load_questions_from_json(json_path: str, language: str = "fr"):
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    return data["questionnaire"]


def seed_questions_and_domains(questions_fr_path: str, questions_en_path: str, guidance_path: str):
    """Charge les questions FR/EN + guidance dans la base (idempotent)."""
    with open(questions_fr_path, encoding="utf-8") as f:
        q_fr = json.load(f)["questionnaire"]
    with open(questions_en_path, encoding="utf-8") as f:
        q_en = {q["id"]: q for q in json.load(f)["questionnaire"]}
    with open(guidance_path, encoding="utf-8") as f:
        guidance = json.load(f)["guidance"]

    domain_id_map = {
        "Gouvernance": "dom_gov",
        "Accès & Identités": "dom_acc",
        "Infrastructure & Sécurité réseau": "dom_infra",
        "Incidents & Continuité": "dom_inc",
        "Sensibilisation & Formation": "dom_sens",
    }

    with get_connection() as conn:
        for q in q_fr:
            en = q_en.get(q["id"], {})
            gid = guidance.get(q["id"], {})
            conn.execute(
                """INSERT OR REPLACE INTO question
                   (id, domain_id, display_order, text_fr, text_en, guidance_fr, guidance_en)
                   VALUES (?, ?, ?, ?, ?, ?, ?)""",
                (
                    q["id"],
                    domain_id_map.get(q["domain"], "dom_gov"),
                    q["order"],
                    q["text"],
                    en.get("text", q["text"]),
                    gid.get("fr", ""),
                    gid.get("en", ""),
                ),
            )
    return len(q_fr)


def get_all_questions():
    with get_connection() as conn:
        rows = conn.execute(
            """SELECT q.*, d.name_fr as domain_name_fr, d.name_en as domain_name_en
               FROM question q JOIN domain d ON q.domain_id = d.id
               ORDER BY q.display_order"""
        ).fetchall()
        return [dict(r) for r in rows]


def get_all_domains():
    with get_connection() as conn:
        rows = conn.execute("SELECT * FROM domain ORDER BY display_order").fetchall()
        return [dict(r) for r in rows]


# ============================================================
# Recommendations
# ============================================================

def seed_recommendations_from_xlsx(xlsx_path: str):
    """Charge les recommandations depuis le fichier Excel (une feuille par domaine)."""
    import openpyxl

    domain_sheet_map = {
        "Gouvernance": "dom_gov",
        "Acces & Identites": "dom_acc",
        "Infrastructure": "dom_infra",
        "Incidents & Continuite": "dom_inc",
        "Sensibilisation": "dom_sens",
    }
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    count = 0
    with get_connection() as conn:
        for sheet_name, domain_id in domain_sheet_map.items():
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=4, values_only=True):
                if not row or row[0] is None:
                    continue
                rec_id, severity, priority, text_fr, text_en, reference, effort = (list(row) + [None] * 7)[:7]
                conn.execute(
                    """INSERT OR REPLACE INTO recommendation
                       (id, domain_id, severity, priority, text_fr, text_en, reference, effort)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                    (rec_id, domain_id, severity, priority, text_fr, text_en or text_fr, reference, effort),
                )
                count += 1
    return count


def get_recommendations_for_domain(domain_id: str, severities: list):
    placeholders = ",".join("?" * len(severities))
    with get_connection() as conn:
        rows = conn.execute(
            f"""SELECT * FROM recommendation WHERE domain_id = ? AND severity IN ({placeholders})
                ORDER BY priority ASC""",
            (domain_id, *severities),
        ).fetchall()
        return [dict(r) for r in rows]


# ============================================================
# Audits
# ============================================================

def create_audit(pme_id: str) -> str:
    audit_id = new_uuid()
    with get_connection() as conn:
        conn.execute(
            "INSERT INTO audit (id, pme_id, questionnaire_version) VALUES (?, ?, '1.0')",
            (audit_id, pme_id),
        )
    return audit_id


def save_response(audit_id: str, question_id: str, response: int, domain_score: float = None):
    with get_connection() as conn:
        conn.execute(
            "INSERT INTO audit_response (id, audit_id, question_id, response, domain_score) VALUES (?, ?, ?, ?, ?)",
            (new_uuid(), audit_id, question_id, response, domain_score),
        )


def complete_audit(audit_id: str, global_score: float):
    with get_connection() as conn:
        conn.execute(
            "UPDATE audit SET global_score = ?, completed_at = ? WHERE id = ?",
            (global_score, datetime.utcnow().isoformat(), audit_id),
        )


def get_audit_history(pme_id: str):
    with get_connection() as conn:
        rows = conn.execute(
            """SELECT * FROM audit WHERE pme_id = ? AND completed_at IS NOT NULL
               ORDER BY completed_at DESC""",
            (pme_id,),
        ).fetchall()
        return [dict(r) for r in rows]


def get_audit_responses(audit_id: str):
    with get_connection() as conn:
        rows = conn.execute(
            """SELECT ar.*, q.text_fr, q.domain_id FROM audit_response ar
               JOIN question q ON ar.question_id = q.id
               WHERE ar.audit_id = ? ORDER BY q.display_order""",
            (audit_id,),
        ).fetchall()
        return [dict(r) for r in rows]


def log_action(pme_id: str, action: str, details: str = ""):
    with get_connection() as conn:
        conn.execute(
            "INSERT INTO audit_log (id, pme_id, action, details) VALUES (?, ?, ?, ?)",
            (new_uuid(), pme_id, action, details),
        )
